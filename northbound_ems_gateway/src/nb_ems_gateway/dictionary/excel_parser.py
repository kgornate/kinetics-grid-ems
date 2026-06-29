from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from nb_ems_gateway.dictionary.register_map import RegisterMap
from nb_ems_gateway.dictionary.register_point import RegisterPoint
from nb_ems_gateway.normalization.section_mapper import entity_to_asset_id, entity_to_poll_group
from nb_ems_gateway.normalization.signal_name_mapper import normalize_signal_name


HEADER_ROW = 2


def parse_excel_register_map(path: str | Path, *, name: str = "china_ems_northbound", version: str = "v1") -> RegisterMap:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(v).strip() if v is not None else "" for v in next(sheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))]
    index = {header: idx for idx, header in enumerate(headers)}
    required = [
        "Channel Name", "Port No.", "Device ID", "Register Addr.", "Register Qty.",
        "Group No.", "Entity Name", "Point Name", "Point Type", "Unit", "Description",
        "R/W (0: RO, 1: RW)", "Factor",
    ]
    missing = [h for h in required if h not in index]
    if missing:
        raise ValueError(f"Missing expected columns: {missing}")

    points: list[RegisterPoint] = []
    for excel_row_no, row in enumerate(sheet.iter_rows(min_row=HEADER_ROW + 1, values_only=True), start=HEADER_ROW + 1):
        if not any(cell is not None for cell in row):
            continue
        entity_name = str(_cell(row, index, "Entity Name") or "").strip()
        point_name = str(_cell(row, index, "Point Name") or "").strip()
        if not entity_name or not point_name:
            continue
        address = _int(_cell(row, index, "Register Addr."))
        register_qty = _int(_cell(row, index, "Register Qty."))
        asset_id = entity_to_asset_id(entity_name)
        point_id = f"{asset_id}_{address:05d}"
        unit = _optional_str(_cell(row, index, "Unit"))
        points.append(
            RegisterPoint(
                point_id=point_id,
                channel_name=_optional_str(_cell(row, index, "Channel Name")),
                port=_optional_int(_cell(row, index, "Port No.")),
                unit_id=_optional_int(_cell(row, index, "Device ID")),
                address=address,
                register_qty=register_qty,
                group_no=_optional_int(_cell(row, index, "Group No.")),
                entity_name=entity_name,
                point_name=point_name,
                point_type=str(_cell(row, index, "Point Type") or "").strip(),
                unit=unit,
                description=_optional_str(_cell(row, index, "Description")),
                rw_flag=_int(_cell(row, index, "R/W (0: RO, 1: RW)")),
                factor=_float(_cell(row, index, "Factor"), default=1.0),
                software_access="read_only",
                normalized_name=normalize_signal_name(asset_id, point_name, unit),
                asset_id=asset_id,
                poll_group=entity_to_poll_group(entity_name, point_name),
            )
        )
    return RegisterMap.from_points(name=name, version=version, source_file=str(path), points=points)


def _cell(row: tuple[Any, ...], index: dict[str, int], name: str) -> Any:
    idx = index[name]
    return row[idx] if idx < len(row) else None


def _optional_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _optional_int(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    return int(float(str(value)))


def _int(value: Any) -> int:
    if value is None or str(value).strip() == "":
        raise ValueError("Expected integer cell value")
    return int(float(str(value)))


def _float(value: Any, default: float = 1.0) -> float:
    if value is None or str(value).strip() == "":
        return default
    return float(str(value))
