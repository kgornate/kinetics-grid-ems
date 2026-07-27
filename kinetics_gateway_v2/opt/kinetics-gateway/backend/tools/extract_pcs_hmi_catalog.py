from __future__ import annotations

import argparse
import json
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ADDRESS_RE = re.compile(r"0x[0-9A-Fa-f]{1,4}")

KEYS: dict[int, tuple[str, str]] = {
    0x1100: ("dc_bus_voltage", "DC bus voltage"),
    0x1101: ("dc_bus_current", "DC bus current"),
    0x1102: ("battery_voltage", "Battery voltage"),
    0x1103: ("battery_current", "Battery current"),
    0x1104: ("dc_power", "DC power"),
    0x1105: ("grid_ab_voltage", "Grid AB line voltage"),
    0x1106: ("grid_bc_voltage", "Grid BC line voltage"),
    0x1107: ("grid_ca_voltage", "Grid CA line voltage"),
    0x1108: ("grid_a_current", "Grid phase A current"),
    0x1109: ("grid_b_current", "Grid phase B current"),
    0x110A: ("grid_c_current", "Grid phase C current"),
    0x110B: ("power_factor", "Power factor"),
    0x110C: ("grid_frequency", "Grid frequency"),
    0x110D: ("grid_active_power", "Grid active power"),
    0x110E: ("grid_reactive_power", "Grid reactive power"),
    0x110F: ("grid_apparent_power", "Grid apparent power"),
    0x1110: ("igbt_a_temperature", "IGBT A temperature"),
    0x1111: ("igbt_b_temperature", "IGBT B temperature"),
    0x1112: ("igbt_c_temperature", "IGBT C temperature"),
    0x1113: ("cabinet_temperature", "Cabinet internal temperature"),
    0x1114: ("positive_bus_voltage", "Positive DC bus voltage"),
    0x1115: ("negative_bus_voltage", "Negative DC bus voltage"),
    0x1116: ("inverter_ab_voltage", "Inverter AB line voltage"),
    0x1117: ("inverter_bc_voltage", "Inverter BC line voltage"),
    0x1118: ("inverter_ca_voltage", "Inverter CA line voltage"),
    0x1119: ("pcc_ab_voltage", "PCC AB line voltage"),
    0x111A: ("auxiliary_bus_voltage", "Auxiliary bus voltage"),
    0x111B: ("pcc_phase_compensation", "PCC phase compensation"),
    0x111C: ("can_module_count", "CAN communication module count"),
    0x111D: ("phase_a_to_ground_voltage", "Phase A to ground voltage"),
    0x111E: ("phase_b_to_ground_voltage", "Phase B to ground voltage"),
    0x111F: ("phase_c_to_ground_voltage", "Phase C to ground voltage"),
    0x1120: ("battery_positive_ground_impedance", "Battery positive-to-ground impedance"),
    0x1121: ("battery_negative_ground_impedance", "Battery negative-to-ground impedance"),
    0x112B: ("grid_n_current", "Grid neutral current"),
    0x1200: ("operating_state", "PCS operating state"),
    0x1201: ("status_word_1", "PCS status word 1"),
    0x1202: ("status_word_2", "PCS status word 2"),
    0x1203: ("status_word_3", "PCS status word 3"),
    0x1204: ("actual_product_mode", "Actual product mode"),
    0x1205: ("actual_pq_mode", "Actual PQ mode"),
    0x120A: ("can_receive_life", "CAN receive life counter"),
}

DATA_TYPES = {"0": "S16", "1": "U16", "2": "U16", "3": "U16", "4": "UNKNOWN", "5": "UNKNOWN", "6": "UNKNOWN", "16": "UNKNOWN"}


def load_json_gbk(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="gbk"))


def parse_address(value: Any) -> int | None:
    match = ADDRESS_RE.fullmatch(str(value or "").strip())
    return int(match.group(0), 16) if match else None


def parse_number(value: Any) -> float | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_int(value: Any, default: int = 1) -> int:
    number = parse_number(value)
    return int(number) if number is not None else default


def parse_enum(text: str) -> dict[str, str]:
    result: dict[str, str] = {}
    for item in str(text or "").split("|"):
        item = item.strip()
        match = re.match(r"^(-?\d+)\s+(.+)$", item)
        if match:
            result[match.group(1)] = match.group(2).strip()
    return result


def category_for(address: int) -> str:
    if 0x1100 <= address <= 0x11FF:
        return "measure"
    if 0x1200 <= address <= 0x12FF:
        return "status"
    if 0x1300 <= address <= 0x13FF:
        return "version"
    if 0x1400 <= address <= 0x14FF:
        return "parameter"
    if 0x1500 <= address <= 0x15FF:
        return "control_parameter"
    if 0x1600 <= address <= 0x16FF:
        return "protection_parameter"
    return "raw"


def poll_class_for(category: str) -> str:
    return {"measure": "normal", "status": "normal", "version": "slow"}.get(category, "slow")


def candidate_rank(address: int, menu: str, page: str) -> int:
    if address >= 0x1400 and menu == "PCS参数设置":
        return 100
    if address < 0x1400 and page == "PCS运行数据":
        return 100
    if page == "总览":
        return 80
    return 50


def parse_cfg_points(cfg: dict[str, Any]) -> dict[int, dict[str, Any]]:
    candidates: dict[int, list[dict[str, Any]]] = {}
    for conf in cfg.get("configure", []):
        for menu in conf.get("main_menu_button_names", []):
            menu_name = str(menu.get("main_menu_name") or "")
            for page in menu.get("main_menu_details", []):
                page_name = str(page.get("navigation_pages_names") or "")
                for row in page.get("navigation_details", []):
                    if not isinstance(row, list) or len(row) < 15:
                        continue
                    address = parse_address(row[0])
                    if address is None:
                        continue
                    signal_type = str(row[2]).strip()
                    data_code = str(row[3]).strip()
                    if signal_type not in {"0", "1", "2", "3", "4", "5", "6"}:
                        continue
                    if data_code not in DATA_TYPES:
                        continue
                    scale = parse_number(row[9])
                    register_count = max(1, parse_int(row[13], 1))
                    candidate = {
                        "address": address,
                        "name_cn": str(row[1]).strip(),
                        "signal_type": signal_type,
                        "data_type": DATA_TYPES[data_code],
                        "data_type_code": data_code,
                        "enum_text_cn": str(row[4]).strip() or None,
                        "enum": parse_enum(str(row[4])),
                        "scale": scale,
                        "minimum": parse_number(row[10]),
                        "maximum": parse_number(row[11]),
                        "register_count": register_count,
                        "unit": str(row[14]).strip() or None,
                        "menu": menu_name,
                        "page": page_name,
                        "rank": candidate_rank(address, menu_name, page_name),
                        "source": "cfg_sys.json",
                    }
                    candidates.setdefault(address, []).append(candidate)
    selected: dict[int, dict[str, Any]] = {}
    for address, items in candidates.items():
        selected[address] = sorted(items, key=lambda item: item["rank"], reverse=True)[0]
    return selected


def parse_alarm_bitfields(cfg_alarm: dict[str, Any]) -> dict[int, list[dict[str, Any]]]:
    output: dict[int, list[dict[str, Any]]] = {}
    for conf in cfg_alarm.get("configure", []):
        for item in conf.get("main_alarm_cfg_items", []):
            address = parse_address(item.get("fault_register"))
            if address is None or not (0x1210 <= address <= 0x122A):
                continue
            fields: list[dict[str, Any]] = []
            for mask_text, name_cn in item.get("fault_content", []):
                try:
                    mask = int(str(mask_text), 16)
                except ValueError:
                    continue
                if mask <= 0 or mask & (mask - 1):
                    continue
                bit = int(math.log2(mask))
                fields.append({
                    "bit": bit,
                    "key": f"fault_bit_{bit}",
                    "name_cn": str(name_cn).strip(),
                })
            if fields:
                output[address] = fields
    return output


def build_catalog(pcs_xlsx: Path, cfg_sys_path: Path, cfg_alarm_path: Path) -> dict[str, Any]:
    ws = load_workbook(pcs_xlsx, data_only=True)["PCS"]
    workbook: dict[int, dict[str, Any]] = {}
    for row_index in range(2, ws.max_row + 1):
        address = parse_address(ws.cell(row_index, 2).value)
        if address is None:
            continue
        workbook[address] = {
            "serial": ws.cell(row_index, 1).value,
            "name_cn": str(ws.cell(row_index, 3).value or "").strip(),
            "sample_frequency": ws.cell(row_index, 4).value,
            "function_code": int(ws.cell(row_index, 5).value or 3),
            "notes": ws.cell(row_index, 6).value,
            "source_row": row_index,
        }

    cfg_sys = load_json_gbk(cfg_sys_path)
    cfg_points = parse_cfg_points(cfg_sys)
    alarm_fields = parse_alarm_bitfields(load_json_gbk(cfg_alarm_path))
    addresses = sorted(set(workbook) | set(cfg_points))
    points: list[dict[str, Any]] = []

    for address in addresses:
        base = workbook.get(address, {})
        cfg = cfg_points.get(address, {})
        category = category_for(address)
        key, name_en = KEYS.get(address, (f"reg_{address:04x}", None))
        data_type = cfg.get("data_type") or ("UNKNOWN" if category == "measure" else "U16")
        register_count = int(cfg.get("register_count") or 1)
        point = {
            "id": f"pcs.{category}.{key}",
            "sheet": "PCS",
            "scope": "pcs",
            "category": category,
            "serial": base.get("serial"),
            "address": address,
            "address_hex": f"0x{address:04X}",
            "key": key,
            "name_cn": base.get("name_cn") or cfg.get("name_cn"),
            "name_en": name_en,
            "sample_frequency": base.get("sample_frequency", 5),
            "function_code": int(base.get("function_code") or 3),
            "read_function": int(base.get("function_code") or 3),
            "access": "R",
            "write_candidate": cfg.get("signal_type") == "3",
            "data_type": data_type,
            "register_width": 1,
            "element_count": register_count,
            "register_count": register_count,
            "scale": cfg.get("scale"),
            "unit": cfg.get("unit"),
            "range_text": (
                f"{cfg.get('minimum')}..{cfg.get('maximum')}"
                if cfg.get("minimum") is not None and cfg.get("maximum") is not None
                else None
            ),
            "enum_text_cn": cfg.get("enum_text_cn"),
            "enum": cfg.get("enum") or {},
            "bitfields": alarm_fields.get(address, []),
            "poll_class": poll_class_for(category),
            "poll_enabled": True,
            "reserved": "预留" in str(base.get("name_cn") or cfg.get("name_cn") or ""),
            "hardware_validation": (
                "validated_2026-07-27"
                if (0x1100 <= address <= 0x1121) or address == 0x1200
                else "hmi_map_not_yet_directly_probed"
            ),
            "source": (
                "PCS.xlsx + cfg_sys.json"
                if address in workbook and address in cfg_points
                else "PCS.xlsx" if address in workbook else "cfg_sys.json"
            ),
            "source_row": base.get("source_row"),
            "notes": base.get("notes"),
        }
        points.append(point)

    return {
        "metadata": {
            "source": "Working PCS Windows HMI package supplied 2026-07-27",
            "workbook": pcs_xlsx.name,
            "cfg_sys": cfg_sys_path.name,
            "cfg_alarm": cfg_alarm_path.name,
            "point_count": len(points),
            "workbook_point_count": len(workbook),
            "hmi_cfg_unique_point_count": len(cfg_points),
            "confirmed_transport": "Modbus RTU over RS485",
            "confirmed_serial": {"baudrate": 38400, "bytesize": 8, "parity": "N", "stopbits": 1},
            "confirmed_unit_ids": [1],
            "planned_device_count": 4,
            "writes_enabled": False,
            "hardware_validation": {
                "unit_id": 1,
                "function_code": 3,
                "validated_ranges": ["0x1100-0x1121", "0x1200"],
                "stability_reads": 10,
                "stability_successes": 10,
            },
        },
        "points": points,
        "reserved_ranges": [],
        "overrides_schema": {
            "points": {
                "0x1100": {"data_type": "S16", "scale": 0.1, "unit": "V"}
            }
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate PCS catalog from the working HMI files")
    parser.add_argument("--pcs-xlsx", required=True)
    parser.add_argument("--cfg-sys", required=True)
    parser.add_argument("--cfg-alarm", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    catalog = build_catalog(Path(args.pcs_xlsx), Path(args.cfg_sys), Path(args.cfg_alarm))
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(catalog["metadata"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
