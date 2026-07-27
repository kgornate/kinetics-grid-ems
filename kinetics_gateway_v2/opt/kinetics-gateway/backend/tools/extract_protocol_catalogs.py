from __future__ import annotations

import argparse
import json
import re
import unicodedata
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


STANDARD_BMS_SHEETS = {
    "Bank Signal": ("bank", "signal", 4, "fast"),
    "Bank Measure": ("bank", "measure", 4, "fast"),
    "Bank Control": ("bank", "control", 3, "slow"),
    "Bank Parameter Set & Read": ("bank", "parameter", 3, "slow"),
    "Rack Signal": ("rack", "signal", 4, "fast"),
    "Rack Detail": ("rack", "detail", 4, "bulk"),
    "Rack Measure": ("rack", "measure", 4, "normal"),
    "Rack Control": ("rack", "control", 3, "slow"),
    "Rack_Parameter Set & Read": ("rack", "parameter", 3, "slow"),
    "Rack_Alarm_lv4_Para Set& Read": ("rack", "parameter_lv4", 3, "slow"),
    "Env(signal)": ("environment", "signal", 4, "normal"),
}


@dataclass
class BitField:
    bit: int
    key: str
    name_en: str | None = None
    name_cn: str | None = None
    enum_text_en: str | None = None
    enum_text_cn: str | None = None


@dataclass
class RegisterPoint:
    id: str
    sheet: str
    scope: str
    category: str
    address: int
    address_hex: str
    key: str
    name_en: str | None
    name_cn: str | None
    access: str
    data_type: str
    element_count: int
    register_width: int
    register_count: int
    scale: float | None
    unit: str | None
    range_text: str | None
    enum_text_en: str | None
    enum_text_cn: str | None
    usage_en: str | None
    usage_cn: str | None
    source: str | None
    notes: str | None
    read_function: int
    write_functions: list[int]
    poll_class: str
    reserved: bool
    bitfields: list[BitField] = field(default_factory=list)
    source_row: int | None = None


HEX_RE = re.compile(r"0x[0-9a-fA-F]+")
BIT_RE = re.compile(r"bit\s*(\d+)", re.IGNORECASE)


def slugify(value: Any, fallback: str) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^A-Za-z0-9_]+", "_", text).strip("_").lower()
    text = re.sub(r"_+", "_", text)
    if not text or text[0].isdigit():
        text = f"{fallback}_{text}" if text else fallback
    return text


def parse_hex(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    match = HEX_RE.search(text)
    if match:
        return int(match.group(0), 16)
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def parse_int(value: Any, default: int = 1) -> int:
    if value in (None, ""):
        return default
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    try:
        return int(text, 0)
    except ValueError:
        try:
            return int(float(text))
        except ValueError:
            return default


def normalize_type(value: Any) -> str:
    text = str(value or "U16").strip().upper()
    aliases = {"UINT16": "U16", "INT16": "S16", "UINT32": "U32", "INT32": "S32"}
    return aliases.get(text, text)


def width_for_type(data_type: str) -> int:
    match = re.search(r"(16|32|64)$", data_type)
    if not match:
        return 1
    return int(match.group(1)) // 16


def get_version(workbook) -> dict[str, Any]:
    ws = workbook["Version"]
    latest: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] and row[1]:
            latest = {
                "date": row[0].isoformat() if hasattr(row[0], "isoformat") else str(row[0]),
                "version": str(row[1]),
                "description": str(row[2] or ""),
            }
    return latest


def find_header_row(ws, expected: str = "Modbus地址") -> tuple[int, list[str]]:
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        values = [str(v).strip() if v is not None else "" for v in row]
        if expected in values:
            return row_idx, values
    raise ValueError(f"Could not find header '{expected}' in {ws.title}")


def value(row: tuple[Any, ...], mapping: dict[str, int], name: str) -> Any:
    idx = mapping.get(name)
    return row[idx] if idx is not None and idx < len(row) else None


def parse_bit(text: Any) -> int | None:
    if text is None:
        return None
    match = BIT_RE.search(str(text))
    return int(match.group(1)) if match else None


def parse_standard_sheet(ws, scope: str, category: str, read_function: int, poll_class: str) -> list[RegisterPoint]:
    header_row, header = find_header_row(ws)
    mapping = {name: i for i, name in enumerate(header) if name}
    points: list[RegisterPoint] = []
    current: RegisterPoint | None = None
    duplicate_counter: dict[str, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        address = parse_hex(value(row, mapping, "Modbus地址"))
        if address is None and ws.title == "Env(signal)":
            # Some cached files expose the numerical address in an unlabeled column.
            for cell in reversed(row):
                candidate = parse_hex(cell)
                if candidate is not None and candidate >= 0x0600:
                    address = candidate
                    break

        bit = parse_bit(value(row, mapping, "英文使用说明"))
        if bit is None:
            bit = parse_bit(value(row, mapping, "中文使用说明"))

        if address is None:
            if current is not None and bit is not None:
                child_key = slugify(value(row, mapping, "英文子key"), f"bit_{bit}")
                current.bitfields.append(
                    BitField(
                        bit=bit,
                        key=child_key,
                        name_en=str(value(row, mapping, "英文使用说明") or "").strip() or None,
                        name_cn=str(value(row, mapping, "中文使用说明") or "").strip() or None,
                        enum_text_en=str(value(row, mapping, "说明(值/枚举说明)(英文)") or "").strip() or None,
                        enum_text_cn=str(value(row, mapping, "说明(值/枚举说明)") or "").strip() or None,
                    )
                )
            continue

        name_en = str(value(row, mapping, "配置项英文名") or "").strip() or None
        name_cn = str(value(row, mapping, "配置项中文名") or "").strip() or None
        key_source = value(row, mapping, "英文key") or name_en or name_cn
        key = slugify(key_source, f"reg_{address:04x}")
        base_id = f"{scope}.{category}.{key}"
        duplicate_counter[base_id] = duplicate_counter.get(base_id, 0) + 1
        point_id = base_id if duplicate_counter[base_id] == 1 else f"{base_id}_{duplicate_counter[base_id]}"

        data_type = normalize_type(value(row, mapping, "数值格式"))
        element_count = max(1, parse_int(value(row, mapping, "个数"), 1))
        register_width = width_for_type(data_type)
        access = str(value(row, mapping, "是否可写") or "R").strip().upper()
        scale_raw = value(row, mapping, "单位系数")
        scale: float | None
        try:
            scale = float(scale_raw) if scale_raw not in (None, "") else None
        except (TypeError, ValueError):
            scale = None
        reserved = "reserve" in key or "预留" in str(name_cn or "")
        write_functions = [6, 16] if "W" in access else []

        current = RegisterPoint(
            id=point_id,
            sheet=ws.title,
            scope=scope,
            category=category,
            address=address,
            address_hex=f"0x{address:04X}",
            key=key,
            name_en=name_en,
            name_cn=name_cn,
            access=access,
            data_type=data_type,
            element_count=element_count,
            register_width=register_width,
            register_count=element_count * register_width,
            scale=scale,
            unit=str(value(row, mapping, "单位") or "").strip() or None,
            range_text=str(value(row, mapping, "范围") or "").strip() or None,
            enum_text_en=str(value(row, mapping, "说明(值/枚举说明)(英文)") or "").strip() or None,
            enum_text_cn=str(value(row, mapping, "说明(值/枚举说明)") or "").strip() or None,
            usage_en=str(value(row, mapping, "英文使用说明") or "").strip() or None,
            usage_cn=str(value(row, mapping, "中文使用说明") or "").strip() or None,
            source=str(value(row, mapping, "数据源") or "").strip() or None,
            notes=str(value(row, mapping, "备注") or "").strip() or None,
            read_function=read_function,
            write_functions=write_functions,
            poll_class="disabled" if reserved else poll_class,
            reserved=reserved,
            source_row=row_idx,
        )
        if bit is not None:
            current.bitfields.append(
                BitField(
                    bit=bit,
                    key=slugify(value(row, mapping, "英文子key"), f"bit_{bit}"),
                    name_en=str(value(row, mapping, "英文使用说明") or "").strip() or None,
                    name_cn=str(value(row, mapping, "中文使用说明") or "").strip() or None,
                    enum_text_en=current.enum_text_en,
                    enum_text_cn=current.enum_text_cn,
                )
            )
        points.append(current)
    return points


def parse_range_sheet(ws, scope: str, category: str) -> list[dict[str, Any]]:
    header_row, header = find_header_row(ws)
    mapping = {name: i for i, name in enumerate(header) if name}
    ranges: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        address_text = value(row, mapping, "Modbus地址")
        if not address_text:
            continue
        matches = HEX_RE.findall(str(address_text))
        if not matches:
            continue
        start = int(matches[0], 16)
        end = int(matches[-1], 16) if len(matches) > 1 else start
        name_en = value(row, mapping, "配置项英文名")
        name_cn = value(row, mapping, "配置项中文名")
        key = slugify(name_en or name_cn, f"range_{start:04x}")
        ranges.append(
            {
                "id": f"{scope}.{category}.{key}",
                "sheet": ws.title,
                "scope": scope,
                "category": category,
                "key": key,
                "name_en": str(name_en or "").strip() or None,
                "name_cn": str(name_cn or "").strip() or None,
                "address_start": start,
                "address_end": end,
                "address_range": f"0x{start:04X}-0x{end:04X}",
                "register_count": end - start + 1,
                "access": str(value(row, mapping, "是否可写") or "R/W").strip().upper(),
                "data_type": normalize_type(value(row, mapping, "数值格式")),
                "source": str(value(row, mapping, "数据源") or "").strip() or None,
                "notes": str(value(row, mapping, "中文使用说明") or "").strip() or None,
                "poll_class": "disabled",
                "reserved": True,
                "source_row": row_idx,
            }
        )
    return ranges


def extract_bms(input_path: Path) -> dict[str, Any]:
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    points: list[RegisterPoint] = []
    for sheet, (scope, category, read_function, poll_class) in STANDARD_BMS_SHEETS.items():
        points.extend(parse_standard_sheet(workbook[sheet], scope, category, read_function, poll_class))

    ranges = []
    ranges.extend(parse_range_sheet(workbook["Env(control)"], "environment", "control_range"))
    ranges.extend(parse_range_sheet(workbook["Env(default)"], "environment", "device_range"))

    counts: dict[str, int] = {}
    for point in points:
        key = f"{point.scope}.{point.category}"
        counts[key] = counts.get(key, 0) + 1

    return {
        "metadata": {
            "source_file": input_path.name,
            "workbook_version": get_version(workbook),
            "protocols": ["modbus_tcp", "modbus_rtu", "vendor_can_register_protocol"],
            "byte_order": "big",
            "word_order": "big",
            "supported_modbus_functions": [3, 4, 6, 16],
            "recommended_max_read_registers": 126,
            "default_modbus_tcp": {
                "host": "10.30.4.13",
                "three_level_bau_port": 503,
                "bau_unit_id": 1,
                "bcu_unit_id_rule": "2 to N+1",
                "power_environment_unit_id": 127,
            },
            "read_function_note": "Read function assignments are inferred by sheet semantics: telemetry sheets prefer FC04; control/parameter sheets prefer FC03. The runtime can retry the alternate read function and allows per-point overrides.",
        },
        "counts": counts,
        "points": [
            {
                **{k: v for k, v in asdict(point).items() if k != "bitfields"},
                "bitfields": [asdict(bit) for bit in point.bitfields],
            }
            for point in points
        ],
        "reserved_ranges": ranges,
    }


def extract_pcs(input_path: Path) -> dict[str, Any]:
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    ws = workbook[workbook.sheetnames[0]]
    rows: list[dict[str, Any]] = []
    seen: dict[str, int] = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        serial, address_value, name_cn, sample_frequency, function_code, notes = row[:6]
        address = parse_hex(address_value)
        if address is None:
            continue
        key = slugify(name_cn, f"reg_{address:04x}")
        seen[key] = seen.get(key, 0) + 1
        if seen[key] > 1:
            key = f"{key}_{seen[key]}"
        rows.append(
            {
                "id": f"pcs.raw.{key}",
                "sheet": ws.title,
                "scope": "pcs",
                "category": "raw",
                "serial": serial,
                "address": address,
                "address_hex": f"0x{address:04X}",
                "key": key,
                "name_cn": str(name_cn or "").strip() or None,
                "name_en": None,
                "sample_frequency": sample_frequency,
                "function_code": int(function_code or 3),
                "access": "R",
                "data_type": "UNKNOWN",
                "register_count": 1,
                "scale": None,
                "unit": None,
                "byte_order": "UNKNOWN",
                "word_order": "UNKNOWN",
                "notes": str(notes or "").strip() or None,
                "source_row": row_idx,
            }
        )
    return {
        "metadata": {
            "source_file": input_path.name,
            "protocol": "modbus_tcp_expected",
            "register_count": len(rows),
            "limitations": [
                "IP address, TCP port and Unit ID are deployment parameters and are not present.",
                "Data types, scaling, units, enum/bit definitions and word order are not present.",
                "All rows list FC03; writable access, write functions and valid command values are not defined.",
            ],
        },
        "points": rows,
        "overrides_schema": {
            "address_hex": {
                "key": "optional_key",
                "name_en": "optional English name",
                "data_type": "U16|S16|U32|S32|FLOAT32",
                "register_count": 1,
                "scale": 1.0,
                "unit": "engineering unit",
                "access": "R|R/W",
                "write_function": 6,
                "enum": {"0": "Stop", "1": "Run"},
                "bitfields": {"0": "Ready"},
            }
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--bms", type=Path, required=True)
    parser.add_argument("--pcs", type=Path, required=True)
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args()
    args.out.mkdir(parents=True, exist_ok=True)
    bms = extract_bms(args.bms)
    pcs = extract_pcs(args.pcs)
    (args.out / "bms_catalog.json").write_text(json.dumps(bms, indent=2, ensure_ascii=False), encoding="utf-8")
    (args.out / "pcs_catalog.json").write_text(json.dumps(pcs, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"BMS points: {len(bms['points'])}; reserved ranges: {len(bms['reserved_ranges'])}")
    print(f"PCS points: {len(pcs['points'])}")


if __name__ == "__main__":
    main()
